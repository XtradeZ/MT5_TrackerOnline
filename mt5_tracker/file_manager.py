"""
FileManager - Manejo de exportación de archivos

Responsable de leer y escribir archivos CSV/XLSX de forma incremental,
evitando duplicados y manteniendo un archivo único por cuenta.
"""

import os
import csv
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

# Mapa para traducir nombres de columnas del config a claves de datos de trades
# Esto permite que el `config.json` sea más personalizable.
HEADER_MAP = {
    "Ticket": "position_id",
    "Símbolo": "symbol",
    "Tipo": "type",
    "Volumen": "volume",
    "Precio": "price",
    "Beneficio": "profit",
    "Fecha Apertura": "opening_time",
    "Fecha Cierre": "closing_time",
    "Comentario": "comment",
    # Mapeos para los nombres en inglés por si acaso
    "Symbol": "symbol",
    "Type": "type",
    "Volume": "volume",
    "Price": "price",
    "Profit": "profit",
    "Opening Time": "opening_time",
    "Closing Time": "closing_time",
    "Comment": "comment"
}


class FileManager:
    """Maneja la exportación de archivos de forma incremental"""
    
    def __init__(self, config):
        """Inicializa el FileManager con el directorio de salida"""
        self.config = config
        export_settings = self.config.get('export_settings', {})
        self.output_dir = export_settings.get('output_folder', 'output')
        self.columns = export_settings.get('columns', ['Position ID (Ticket)', 'Símbolo', 'Tipo', 'Volumen', 'Precio', 'Beneficio', 'Fecha Apertura', 'Fecha Cierre', 'Comentario'])

        if not os.path.exists(self.output_dir):
            os.makedirs(self.output_dir)
    
    def get_account_file_path(self, account_id, format_type="xlsx"):
        """Genera la ruta del archivo por cuenta"""
        if format_type.lower() == "csv":
            filename = f"Account_{account_id}.csv"
        else:
            filename = f"Account_{account_id}.xlsx"
        return os.path.join(self.output_dir, filename)
    
    def get_existing_operations(self, file_path, format_type="xlsx"):
        """Lee las operaciones existentes del archivo"""
        existing_operations = {}
        
        if not os.path.exists(file_path):
            return existing_operations
        
        try:
            if format_type.lower() == "csv":
                with open(file_path, 'r', encoding='utf-8-sig') as file:
                    reader = csv.DictReader(file)
                    for row in reader:
                        position_id = row.get('Position ID (Ticket)', '')
                        if position_id:
                            existing_operations[position_id] = row
            else:
                wb = load_workbook(file_path)
                ws = wb.active
                
                # Leer encabezados
                headers = []
                for col in range(1, ws.max_column + 1):
                    header_value = ws.cell(row=1, column=col).value
                    if header_value:
                        headers.append(str(header_value))
                
                # Leer datos de forma más segura
                for row in range(2, ws.max_row + 1):
                    # Verificar si la fila tiene datos
                    if ws.cell(row=row, column=1).value is None:
                        continue
                        
                    row_data = {}
                    for col, header in enumerate(headers, 1):
                        cell_value = ws.cell(row=row, column=col).value
                        row_data[header] = str(cell_value) if cell_value is not None else ''
                    
                    position_id = row_data.get('Position ID (Ticket)', '')
                    if position_id and position_id != 'None':
                        existing_operations[position_id] = row_data
                        
        except Exception as e:
            print(f"⚠️ Error al leer archivo existente: {e}")
            
        return existing_operations
    
    def save_to_csv(self, trades_data, account_id):
        """Guarda las operaciones en formato CSV de forma incremental"""
        csv_path = self.get_account_file_path(account_id, "csv")
        
        # Leer operaciones existentes
        existing_operations = self.get_existing_operations(csv_path, "csv")
        
        # Separar operaciones nuevas y operaciones con comentarios actualizados
        new_operations = []
        updated_operations = []
        
        for trade in trades_data:
            position_id = str(trade.get('position_id', ''))
            if position_id:
                if position_id not in existing_operations:
                    # Operación nueva
                    new_operations.append(trade)
                else:
                    # Verificar si el comentario cambió
                    existing_comment = existing_operations[position_id].get('Comentario', '')
                    new_comment = str(trade.get('comment', ''))
                    if existing_comment != new_comment:
                        # Comentario actualizado
                        updated_operations.append(trade)
        
        if not new_operations and not updated_operations:
            print(f"ℹ️ No hay operaciones nuevas o actualizadas para el archivo {csv_path}")
            return csv_path
        
        headers = self.columns
        
        # Si el archivo no existe, crear con encabezados
        file_exists = os.path.exists(csv_path)
        
        try:
            if updated_operations:
                # Si hay operaciones actualizadas, reescribir todo el archivo
                self._rewrite_csv_file(csv_path, headers, trades_data)
                print(f"✅ Archivo actualizado con {len(updated_operations)} operaciones modificadas")
            else:
                # Solo agregar operaciones nuevas
                with open(csv_path, 'a', newline='', encoding='utf-8-sig') as file:
                    writer = csv.writer(file)
                    
                    # Escribir encabezados solo si es archivo nuevo
                    if not file_exists:
                        writer.writerow(headers)
                    
                    # Escribir solo operaciones nuevas
                    for trade in new_operations:
                        row = [str(trade.get(HEADER_MAP.get(h, ''), '')) for h in headers]
                        writer.writerow(row)
                
                print(f"✅ Agregadas {len(new_operations)} operaciones nuevas al archivo {csv_path}")
            
            return csv_path
            
        except PermissionError:
            print(f"❌ ERROR: El archivo {csv_path} está abierto en Excel o en otro programa")
            print(f"💡 SOLUCIÓN: Cierra el archivo y vuelve a exportar")
            raise PermissionError(f"El archivo {csv_path} está abierto. Por favor, ciérralo y vuelve a exportar.")
    
    def _rewrite_csv_file(self, csv_path, headers, trades_data):
        """Reescribe el archivo CSV completo con los datos actualizados"""
        with open(csv_path, 'w', newline='', encoding='utf-8-sig') as file:
            writer = csv.writer(file)
            writer.writerow(headers)
            
            for trade in trades_data:
                row = [str(trade.get(HEADER_MAP.get(h, ''), '')) for h in headers]
                writer.writerow(row)
    
    def save_to_xlsx(self, trades_data, account_id):
        """Guarda las operaciones en formato XLSX de forma incremental"""
        xlsx_path = self.get_account_file_path(account_id, "xlsx")
        
        # Leer operaciones existentes
        existing_operations = self.get_existing_operations(xlsx_path, "xlsx")
        
        # Separar operaciones nuevas y operaciones con comentarios actualizados
        new_operations = []
        updated_operations = []
        
        for trade in trades_data:
            position_id = str(trade.get('position_id', ''))
            if position_id:
                if position_id not in existing_operations:
                    # Operación nueva
                    new_operations.append(trade)
                else:
                    # Verificar si el comentario cambió
                    existing_comment = existing_operations[position_id].get('Comentario', '')
                    new_comment = str(trade.get('comment', ''))
                    if existing_comment != new_comment:
                        # Comentario actualizado
                        updated_operations.append(trade)
        
        if not new_operations and not updated_operations:
            print(f"ℹ️ No hay operaciones nuevas o actualizadas para el archivo {xlsx_path}")
            return xlsx_path
        
        headers = self.columns
        
        try:
            if updated_operations:
                # Si hay operaciones actualizadas, reescribir todo el archivo
                self._rewrite_xlsx_file(xlsx_path, headers, trades_data, account_id)
                print(f"✅ Archivo actualizado con {len(updated_operations)} operaciones modificadas")
            else:
                # Solo agregar operaciones nuevas
                if not os.path.exists(xlsx_path):
                    wb = Workbook()
                    ws = wb.active
                    ws.title = f"Account_{account_id}"
                    
                    # Estilo para encabezados
                    header_font = Font(bold=True, color="FFFFFF")
                    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    
                    # Escribir encabezados
                    for col, header in enumerate(headers, 1):
                        cell = ws.cell(row=1, column=col, value=header)
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = Alignment(horizontal="center")
                    
                    # Escribir operaciones nuevas
                    for row, trade in enumerate(new_operations, 2):
                        for col, header in enumerate(headers, 1):
                            key = HEADER_MAP.get(header, '')
                            ws.cell(row=row, column=col, value=str(trade.get(key, '')))
                    
                    wb.save(xlsx_path)
                else:
                    # Abrir archivo existente y agregar nuevas operaciones
                    wb = load_workbook(xlsx_path)
                    ws = wb.active
                    
                    # Encontrar la última fila con datos de forma más segura
                    last_row = 1
                    for row in range(1, ws.max_row + 1):
                        if ws.cell(row=row, column=1).value is not None:
                            last_row = row
                    
                    # Agregar nuevas operaciones
                    for i, trade in enumerate(new_operations):
                        row_num = last_row + 1 + i
                        for col, header in enumerate(headers, 1):
                            key = HEADER_MAP.get(header, '')
                            ws.cell(row=row_num, column=col, value=str(trade.get(key, '')))
                    
                    wb.save(xlsx_path)
                
                print(f"✅ Agregadas {len(new_operations)} operaciones nuevas al archivo {xlsx_path}")
            
            return xlsx_path
            
        except PermissionError:
            print(f"❌ ERROR: El archivo {xlsx_path} está abierto en Excel")
            print(f"💡 SOLUCIÓN: Cierra Excel y vuelve a exportar")
            raise PermissionError(f"El archivo {xlsx_path} está abierto en Excel. Por favor, ciérralo y vuelve a exportar.")
    
    def _rewrite_xlsx_file(self, xlsx_path, headers, trades_data, account_id):
        """Reescribe el archivo XLSX completo con los datos actualizados"""
        wb = Workbook()
        ws = wb.active
        ws.title = f"Account_{account_id}"
        
        # Estilo para encabezados
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Escribir encabezados
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Escribir todas las operaciones
        for row, trade in enumerate(trades_data, 2):
            for col, header in enumerate(headers, 1):
                ws.cell(row=row, column=col, value=str(trade.get(HEADER_MAP.get(header, ''), '')))
        
        wb.save(xlsx_path)
    
    def open_output_folder(self):
        """Abre la carpeta de salida en el explorador de archivos"""
        try:
            import subprocess
            import platform
            
            if platform.system() == "Windows":
                os.startfile(self.output_dir)
            elif platform.system() == "Darwin":  # macOS
                subprocess.run(["open", self.output_dir])
            else:  # Linux
                subprocess.run(["xdg-open", self.output_dir])
        except Exception as e:
            print(f"⚠️ No se pudo abrir la carpeta: {e}")
            print(f"📁 Carpeta de salida: {os.path.abspath(self.output_dir)}")
